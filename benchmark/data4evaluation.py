from rag import RAG
from camel.agents import ChatAgent
from camel.models import ModelFactory
from camel.types import ModelPlatformType, ModelType
from dotenv import load_dotenv
import os
import json
from openpyxl import Workbook

load_dotenv(dotenv_path='.env')
api_key = os.getenv('GPT_API_KEY')


def init_agent(query):
    agent = ChatAgent(
        system_message=query,
        model=ModelFactory.create(
            model_platform=ModelPlatformType.OPENAI,
            model_type=ModelType.GPT_4O_MINI,
            api_key=api_key,
            model_config_dict={"temperature": 0.4, "max_tokens": 4096},
        )
    )
    return agent

def rag_qa(query):
    agent = init_agent(query)
    retrieved = []
    rag = RAG("vector_data")
    retrieved = rag.rag_retrieve(query)
    user_msg = f"""以下是我为你找到的参考资料，请根据它回答我的问题。

问题：{query}

参考资料：
{str(retrieved)}
"""
    response = agent.step(user_msg)
    return retrieved, response.msg.content


"""
rag_data = []
query_data = ["当前监测显示：DO=3.4 mg/L，比昨日下降了 0.8 单位，pH=8.0，水温 28.5°C。请问是否需要立即调整供氧或循环策略？如果需要，应优先采取哪些操作？", 
            "溶解氧保持在 6.5 mg/L，但池内摄食量下降、水体发白且 DO 夜间波动范围达2.2 mg/L，该情况可能是什么原因导致？下一步应如何应对？还是说这属于正常情况无需采取措施？", 
            "pH 值在过去 48 小时内从 8.1 下降至 7.4，每日下降幅度约 0.3，氨氮为 0.08 mg/L，亚硝酸盐为 0.2 mg/L。请问系统当前是否应干预？如需干预，具体建议操作有哪些？", 
            "pH 稳定在 7.6–7.7，但每天中午值略升至 8.0 后傍晚迅速下降，此时是否需要调整水体碱度？该如何执行？", 
            "当前氨氮=0.32 mg/L，亚硝酸盐=0.27 mg/L，硝酸盐=4 mg/L，pH 为 7.5，DO=5.2mg/L，投喂正常。请问系统应采取哪些操作调整水质？",
            "连续两天亚硝酸盐在 0.5–0.6 mg/L 波动，硝酸盐无明显上升趋势，氨氮为 0.05mg/L。当前的处理方法并未改善数据，下一步应该如何应对以优化水处理流程？", 
            "当前水温从 29°C 骤降至 24°C，虾群摄食减少，但 DO 维持正常（6.1 mg/L），pH 7.9。这种情况应该如何应对？还是说是正常情况无需采取措施？", 
            "过去三天水温在昼夜之间波动超过 4°C，pH 和 DO 在正常范围内，但摄食量明显不稳。请问这种情况下是否应主动调整日间光照/遮阳或其他应对措施？",
            "水体浊度在短时间内（10 分钟）从 1.5 NTU 上升至 18、14.7，再突然升高至 44.8NTU，同时 DO 无明显变化。造成这一浊度骤升的可能原因有哪些？你会采取哪些紧急应对措施？请说明理由。", 
            "池水盐度在 3 天内从 8‰升至 11‰，水位下降 3.5%。请判断是否应立即补水？补水时需注意哪些参数匹配？", 
            "补充淡水后水位恢复至正常，但次日虾群行为异常、摄食量下降。传感器数据显示补水温度与原水温差为 4.1°C。该情况可能是什么原因？下一步应如何应对？"] 

for i, query in enumerate(query_data):
    retrieved, response_rag = rag_qa(query)
    data = {
        "id": i+1,
        "query": query,
        "retrieved": retrieved,
        "output_with_rag": response_rag
    }
    rag_data.append(data)

with open("result4evaluation2.json", "w", encoding="utf-8") as f:
    json.dump(rag_data, f, ensure_ascii=False, indent=2) 
    print("成功写入")

wb = Workbook()
ws = wb.active
ws.title = "result4evaluation2"
ws.cell(row=1, column=1, value="id")
ws.cell(row=1, column=2, value="query")
ws.cell(row=1, column=3, value="retrieved")
ws.cell(row=1, column=4, value="hypothesis")

for i, data in enumerate(rag_data):
    ws.cell(row=i+2, column=1, value=data["id"])
    ws.cell(row=i+2, column=2, value=data["query"])
    ws.cell(row=i+2, column=3, value=str(data["retrieved"]))
    ws.cell(row=i+2, column=4, value=data["output_with_rag"])

wb.save("result4evaluation2.xlsx")
"""

query = "当前氨氮=0.32 mg/L，亚硝酸盐=0.27 mg/L，硝酸盐=4 mg/L，pH 为 7.5，DO=5.2mg/L，投喂正常。请问系统应采取哪些操作调整水质？"
retrieved, response_rag = rag_qa(query)
print(retrieved)
print(response_rag)